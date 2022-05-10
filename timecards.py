import os
import json
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import mysql.connector
import win32com.client as win32
import logging
import datetime

from finbytools import get_table_data, get_planner_perc, get_executor_perc

MASTER_FOLDER = 'Y:\\Shared\\Карточки учета рабочего времени\\Планирование'
CALENDAR_PATH = 'Y:\\Shared\\Карточки учета рабочего времени\\Планирование\\Календарь.xlsm'
MAPPINGS_PATH = 'Y:\\Shared\\Карточки учета рабочего времени\\Планирование\\Справочники.xlsm'
SCRIPTS_FOLDER = 'Y:\\Power BI\\FIN.by Planning\\planning' 
NEW_LINE = '\n'
TAB = '\t'

INSERT_STATEMENT = '''INSERT INTO timecards ( 
    entry_date, 
    planner, 
    executor, 
    project, 
    task, 
    action, 
    comment, 
    hours, 
    perc_of_completion_executor, 
    perc_of_completion_planner,
    entry_type
) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'''

if __name__ ==  '__main__':
    settings_path = f'{SCRIPTS_FOLDER}\\settings.json'
    logs_folder = f'{SCRIPTS_FOLDER}\\logs'
    refresh_start_time = datetime.datetime.now()
    logs_file_name = f'{logs_folder}\\refresh_timecards_{refresh_start_time.strftime(" %m.%d.%Y %H.%M.%S")}.log'
    if not os.path.isdir(logs_folder):
        os.mkdir(logs_folder)

    logging.basicConfig(filename = logs_file_name, level = logging.INFO, format = u'[%(levelname)s] %(asctime)s : %(message)s')

    logFormatter = logging.Formatter(u"[%(levelname)s] %(asctime)s : %(message)s")
    rootLogger = logging.getLogger()

    consoleHandler = logging.StreamHandler()
    consoleHandler.setFormatter(logFormatter)
    rootLogger.addHandler(consoleHandler)

    settings = json.load(open(settings_path, mode='r'))

    database_user = settings['credentials']['db_user']
    database_password = settings['credentials']['password']
    hostname = settings['hostname']
    schema_name = settings['schema']

    weeks_to_load = settings['weeks_no']
    year_to_load = settings['year_no']

    employees = settings['employees']

    send_mail_to = settings['send_mail_to']

    logging.info(f'Settings were loaded from file: {settings_path}')
    logging.info(f'The following periods were specified: {", ".join(list(map(lambda i: "{:02}".format(i), weeks_to_load)))} -> {str(year_to_load)}')
    if len(employees) > 0:
        logging.info(f'The following employees were specified:{(NEW_LINE+TAB)}{(NEW_LINE+TAB).join(employees)}')

    calendar_file = load_workbook(CALENDAR_PATH, data_only=True)
    calendar = get_table_data(calendar_file, 'Calendar', 'calendar_actual')
    calendar['date'] = calendar['date'].dt.date

    max_budget_week = calendar[ calendar['budgetYearNo'] == year_to_load ]['budgetWeekNo'].max()

    logging.info('Calendar has been loaded')

    mappings_file = load_workbook(MAPPINGS_PATH, data_only=True)
    employees_table = get_table_data(mappings_file, 'Справочники', 'Сотрудники')
    all_employees = list(employees_table['Шаблон заголовка'].dropna())

    logging.info('List of all employees has been loaded')

    try:
        connection = mysql.connector.connect(
            host=hostname,
            database=schema_name,
            user=database_user,
            password=database_password
        )

        logging.info(f'[MYSQL CONNECTION ESTABLISHED]: hostname = {hostname}, database = {schema_name}, user = {database_user}')

        cursor = connection.cursor()            

        errors_counter = 0
        total_rows_inserted = 0
        outlook_message_array = []
        errors_messages_array = []

        for week in weeks_to_load:
            week_rows_inserted = 0

            next_week_no = week + 1 if week < max_budget_week else 1
            next_week_year_no = year_to_load if week < max_budget_week else year_to_load + 1

            week_folder_name = str(year_to_load) + '-W' + f'{week:02d}'
            next_week_folder_name = str(next_week_year_no) + '-W' + f'{next_week_no:02d}'
            
            current_week_calendar = calendar[
                (calendar['budgetYearNo'] == year_to_load) & 
                (calendar['budgetWeekNo'] == week)
            ]

            week_start_date = current_week_calendar['date'].min()
            week_end_date = current_week_calendar['date'].max()

            logging.info(f'[{week_folder_name}] Week start date: {week_start_date}, week end date: {week_end_date}')

            timecards_folder = MASTER_FOLDER + '\\Факт\\' + week_folder_name
            is_folder_exists = os.path.isdir(timecards_folder)
            timecards = os.listdir(timecards_folder) if is_folder_exists else []

            logging.info(f'[{week_folder_name}] Timecards folder: {timecards_folder}, folder exists: {is_folder_exists} ')
            logging.info(f'[{week_folder_name}] The following items ({len(timecards)}) are in the folder:{(NEW_LINE+TAB)}{(NEW_LINE+TAB).join(timecards)}')

            next_week_plans_folder = MASTER_FOLDER + '\План\\' + next_week_folder_name
            is_next_week_folder_exists = os.path.isdir(next_week_plans_folder)
            next_week_plans = os.listdir(next_week_plans_folder) if is_next_week_folder_exists else []

            logging.info(f'[{week_folder_name}] Next week plans folder: {next_week_plans_folder}, folder exists: {is_next_week_folder_exists} ')
            logging.info(f'[{week_folder_name}] The following items ({len(next_week_plans)}) are in the folder:{(NEW_LINE+TAB)}{(NEW_LINE+TAB).join(next_week_plans)}')

            plan_vs_actual_dfs = []
            for file_path in next_week_plans:
                plan_path = f'{next_week_plans_folder}\\{file_path}' 
                
                file_name_pattern = f'plan_{next_week_folder_name}_'
                extension = '.xlsm'
                employee = file_path[
                    file_path.find(file_name_pattern) + len(file_name_pattern):
                    file_path.find(extension)
                ]

                if file_path.startswith(file_name_pattern) and file_path.endswith(extension) and (employee in all_employees):
                    plan_file = load_workbook(plan_path, data_only=True)

                    # fnGetPctgOwner

                    source_plan_df = get_table_data(plan_file, 'План-факт PW', 'PlanVsActualPW')

                    plan_vs_actual_df = source_plan_df[
                        [
                            'Постановщик',
                            'Исполнитель',
                            'Проект', 
                            'Задача', 
                            'Часы (факт)', 
                            '% выполнения (Постановщик)'
                        ]
                    ]

                    df_length = len(plan_vs_actual_df)

                    if df_length > 0:
                        plan_vs_actual_dfs.append(plan_vs_actual_df)

                        logging.info(f'[{week_folder_name}] {"[" + file_path + "]" : <35} Rows were added to PlanVsActual dataframe [{str(df_length)}]')
                    
                    else:
                        logging.warning(f'[{week_folder_name}] {"[" + file_path + "]" : <35} File is empty or didn`t pass validation')

            if len(plan_vs_actual_dfs) > 0:
                plan_vs_actual = pd.concat(plan_vs_actual_dfs)
            else:
                plan_vs_actual = pd.DataFrame(
                    {
                        'Дата': pd.Series(dtype='datetime64[ns]').dt.date,
                        'Постановщик': pd.Series(dtype='str'), 
                        'Исполнитель': pd.Series(dtype='str'), 
                        'Проект': pd.Series(dtype='str'), 
                        'Задача': pd.Series(dtype='str'), 
                        'Действие': pd.Series(dtype='str'), 
                        'Комментарий': pd.Series(dtype='str'), 
                        'Часы': pd.Series(dtype='float'),
                        '% выполнения (Исполнитель)': pd.Series(dtype='float'),
                        '% выполнения (Постановщик)': pd.Series(dtype='float')
                    }
                )
            
            logging.info(f'[{week_folder_name}] Total number of rows in the PlanVsActual dataframe: [{str(len(plan_vs_actual))}]')

            for file_path in timecards:
                try:
                    file_name_pattern = f'timecard_{week_folder_name}_'
                    extension = '.xlsm'
                    employee = file_path[
                        file_path.find(file_name_pattern) + len(file_name_pattern):
                        file_path.find(extension)
                    ]

                    if file_path.startswith(file_name_pattern) and (True if len(employees) == 0 else employee in employees) and (employee in all_employees) and file_path.endswith(extension):
                        fact_path = f'{timecards_folder}\\{file_path}' 

                        fact_file = load_workbook(fact_path, data_only=True)

                        # fnGetTimecard

                        # get executor name
                        defined_range_description = fact_file.defined_names['EmployeeName'].attr_text.split('!')
                        sheet_name = defined_range_description[0].replace("'", "")
                        cell_address = defined_range_description[1].replace('$', '')
                        employee_name = fact_file[sheet_name][cell_address].value

                        timecard_df = get_table_data(fact_file, 'Timecard CW', 'Timecard')

                        logging.info(f'[{week_folder_name}] [{file_path}] The file was loaded to the dataframe')

                        timecard_df = timecard_df.drop('Руководитель', axis=1, errors='ignore')
                        timecard_df = timecard_df.melt(id_vars=['Проект', 'Задача', 'Действие', 'Комментарий'],
                                                    var_name='Дата',
                                                    value_name='Часы')
                        timecard_df['Дата'] = pd.to_datetime(
                            timecard_df['Дата'], format='%d.%m.%Y').dt.date
                        timecard_df['Часы'] = pd.to_numeric(
                            timecard_df['Часы'], downcast='float')

                        timecard_df = timecard_df.dropna(subset=['Часы'])

                        df_length = len(timecard_df)

                        logging.info( f'[{week_folder_name}] [{file_path}] {"Entries were unpivoted and typed:" : <60} [{str(df_length)}]' )

                        if df_length > 0:
                            plan_df = get_table_data(fact_file, 'План CW', 'ПланЗадач')
                            plan_df = plan_df[['Проект', 'Задача',
                                            'Постановщик', 'Часы (факт)', '% выполнения (факт)']]

                            # df with tasks that were in progress during report week

                            # plan_df, timecard_df
                            timecard_df[['Постановщик', '% выполнения (Исполнитель)']] = timecard_df.apply(
                                lambda x: pd.Series(get_executor_perc(plan_df, timecard_df, x['Проект'], x['Задача']), 
                                    index=['Постановщик', '% выполнения (Исполнитель)']), axis=1)

                            logging.info(f'[{week_folder_name}] [{file_path}] {"Percent (by executor) was added:" : <60} [{str(len(timecard_df))}]')

                            # df with undone tasks
                            undone_df = plan_df[plan_df['Часы (факт)'] == 0.0]
                            undone_df['Дата'] = week_end_date

                            undone_df = undone_df[['Дата', 'Постановщик',
                                                'Проект', 'Задача', '% выполнения (факт)']]
                            undone_df.rename(columns={'% выполнения (факт)': '% выполнения (Исполнитель)'}, inplace=True)

                            logging.info(f'[{week_folder_name}] [{file_path}] {"Undone tasks were selected:" : <60} [{str(len(undone_df))}]')

                            # combine all entries
                            fact_result_df = pd.concat([timecard_df, undone_df])

                            df_length_before_filter = len(fact_result_df)
                            logging.info(f'[{week_folder_name}] [{file_path}] {"Undone tasks were appended:" : <60} [{str(df_length_before_filter)}]')

                            # fact_result_df = fact_result_df[fact_result_df.apply(lambda x: (not pd.isnull(x['Проект'])) and (
                                # (not pd.isnull(x['Часы'])) or (not pd.isnull(x['% выполнения (Исполнитель)']))), axis=1)]

                            fact_result_df = fact_result_df[
                                (fact_result_df['Проект'].notnull()) &  
                                ( (fact_result_df['Часы'].notnull()) | (fact_result_df['% выполнения (Исполнитель)'].notnull()) )
                            ]

                            df_length_after_filter = len(fact_result_df)

                            logging.info(f'[{week_folder_name}] [{file_path}] {"`Null` rows were filtered out:" : <60} [{df_length_after_filter}]')

                            fact_result_df['Исполнитель'] = employee_name

                            if len(fact_result_df) > 0:
                                fact_result_df['% выполнения (Постановщик)'] = fact_result_df.apply(
                                    lambda row: get_planner_perc(plan_vs_actual, row['Постановщик'], row['Исполнитель'], row['Проект'], row['Задача']), 
                                    axis=1
                                )

                                logging.info(f'[{week_folder_name}] [{file_path}] {"Percent (by planner) was added:" : <60} [{len(fact_result_df)}]')

                                fact_result_df = fact_result_df.replace({np.nan: None})
                                
                                # metadata for selectind and deleting rows before insertion
                                fact_result_df['entry_type'] = 'timecard' 

                                fact_result_df = fact_result_df[
                                    [
                                        'Дата',
                                        'Постановщик', 
                                        'Исполнитель', 
                                        'Проект', 
                                        'Задача', 
                                        'Действие', 
                                        'Комментарий', 
                                        'Часы',
                                        '% выполнения (Исполнитель)',
                                        '% выполнения (Постановщик)',
                                        'entry_type'
                                    ]
                                ]

                                entries_to_load = list(map(tuple, fact_result_df.itertuples(index=False)))

                                logging.info(f'[{week_folder_name}] [{file_path}] {"Dataframe was converted to a list of tuples:" : <60} [{len(entries_to_load)}]')

                                DELETE_STATEMENT = f'''DELETE FROM timecards WHERE 
                                    entry_date >= '{week_start_date}' and 
                                    entry_date <= '{week_end_date}' and 
                                    executor = '{employee_name}' and
                                    entry_type = 'timecard'
                                '''
                                cursor.execute(DELETE_STATEMENT)
                                connection.commit()
    
                                logging.critical(f'[{week_folder_name}] [{file_path}] {"Rows were deleted from the database" : <60} {"[" + str(cursor.rowcount)+ "]": >5}')

                                cursor.executemany(INSERT_STATEMENT, entries_to_load)  
                                connection.commit()

                                result_message = f'[{week_folder_name}] [{file_path}] {"Rows were inserted to the database" : <60} {"[" + str(cursor.rowcount)+ "]": >5}'
                                total_rows_inserted += cursor.rowcount 
                                week_rows_inserted += cursor.rowcount
                                logging.critical(result_message)

                                outlook_message_array.append(result_message)
                            else:
                                result_message = f'[{week_folder_name}] [{file_path}] File is empty or didn`t pass validation'
                                logging.critical(result_message)
                                outlook_message_array.append(result_message)
                
                except Exception as e:
                    result_message = f'[{week_folder_name}] [{file_path}] Exception was raised:{NEW_LINE}{e}'
                    errors_counter += 1
                    errors_messages_array.append(result_message)
                    logging.error(result_message)
                
            if len(plan_vs_actual) > 0:         
                undone_tasks = plan_vs_actual[
                    (plan_vs_actual['Часы (факт)'].isnull()) &
                    (plan_vs_actual['% выполнения (Постановщик)'].notnull())
                ]
                undone_tasks = undone_tasks.drop('Часы (факт)', axis=1, errors='ignore')
                undone_tasks['Дата'] = week_end_date

                undone_tasks[['Действие', 'Комментарий', 'Часы', '% выполнения (Исполнитель)']] = [np.nan, np.nan, np.nan, np.nan] 
                
                undone_tasks['entry_type'] = 'plan_vs_actual'

                undone_tasks = undone_tasks[
                    [
                        'Дата',
                        'Постановщик', 
                        'Исполнитель', 
                        'Проект', 
                        'Задача', 
                        'Действие', 
                        'Комментарий', 
                        'Часы',
                        '% выполнения (Исполнитель)',
                        '% выполнения (Постановщик)',
                        'entry_type'
                    ]
                ]
                
                undone_tasks = undone_tasks.replace({np.nan: None})

                entries_to_load = list(map(tuple, undone_tasks.itertuples(index=False)))

                cursor.execute(f'''DELETE FROM timecards WHERE 
                    entry_date = '{week_end_date}' and
                    hours is null and
                    not perc_of_completion_planner is null and
                    entry_type = 'plan_vs_actual'
                ''')
                connection.commit()

                logging.critical(f'[{week_folder_name}] [PlanVsActual] {"Rows were deleted from the database" : <60} {"[" + str(cursor.rowcount)+ "]": >5}')

                cursor.executemany(INSERT_STATEMENT, entries_to_load)
                connection.commit()

                result_message = f'[{week_folder_name}] [PlanVsActual] {"Rows were inserted to the database" : <60} {"[" + str(cursor.rowcount)+ "]": >5}'
                total_rows_inserted += cursor.rowcount
                week_rows_inserted += cursor.rowcount
                logging.critical(result_message)

                outlook_message_array.append(result_message)

                week_result_message = f'[{week_folder_name}] {"Total number of inserted rows:" : <60} [{week_rows_inserted}]{NEW_LINE}'
                outlook_message_array.append(week_result_message)
                logging.info(week_result_message)

    except mysql.connector.Error as error:
        result_message = f'Exception was raised:{NEW_LINE}{error}'
        errors_counter += 1
        errors_messages_array.append(result_message)
        logging.error(result_message)

    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
            logging.info('MySQL connection is closed')

    refresh_end_time = datetime.datetime.now()
    refresh_duration = refresh_end_time - refresh_start_time

    logging.info(f'Refresh took {refresh_duration.total_seconds()} seconds')

    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = ';'.join(send_mail_to)
    mail.Subject = 'Timecards refresh result'

    mail_message = f'''{'SUCCESS' if errors_counter == 0 else 'FAILURE'}

    Timecards refresh took: {refresh_duration.total_seconds()} seconds.
    Total number of loaded rows: {str(total_rows_inserted)}.
    Total number of errors raised: {str(errors_counter)}.

    Logs file could be seen in the attachements.
    The logs file could be found at: {logs_file_name}

    Detailed info:
    {NEW_LINE.join([message for message in outlook_message_array])}
    {(NEW_LINE + 'Detailed info about errors:' + NEW_LINE + NEW_LINE.join([error for error in errors_messages_array]) if errors_counter > 0 else '')}
    '''
    mail.Body = mail_message
    mail.Attachments.Add(logs_file_name)
    mail.Send()