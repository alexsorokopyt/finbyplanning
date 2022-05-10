import os
import json
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import mysql.connector
import logging
import datetime
import win32com.client as win32

from finbytools import get_table_data

MASTER_FOLDER = 'Y:\\Shared\\Карточки учета рабочего времени\\Планирование'
CALENDAR_PATH = 'Y:\\Shared\\Карточки учета рабочего времени\\Планирование\\Календарь.xlsm'
MAPPINGS_PATH = 'Y:\\Shared\\Карточки учета рабочего времени\\Планирование\\Справочники.xlsm'
CURRENT_FOLDER = 'Y:\\Power BI\\FIN.by Planning\\planning'
NEW_LINE = '\n'
TAB = '\t'


INSERT_STATEMENT = '''INSERT INTO plans ( 
    plan_date, 
    planner, 
    project, 
    task, 
    executor, 
    deadline, 
    is_backlog, 
    planning_period, 
    planned_hours, 
    perc_of_completion 
) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'''

settings_path = f'{CURRENT_FOLDER}\\settings.json'
logs_folder = f'{CURRENT_FOLDER}\\logs'
refresh_start_time = datetime.datetime.now()
logs_file_name = f'{logs_folder}\\refresh_plans_{refresh_start_time.strftime(" %m.%d.%Y %H.%M.%S")}.log'
if not os.path.isdir(logs_folder):
    os.mkdir(logs_folder)
logging.basicConfig(filename = logs_file_name, level = logging.INFO, format = u'[%(levelname)s] %(asctime)s : %(message)s')

logFormatter = logging.Formatter(u"[%(levelname)s] %(asctime)s : %(message)s")
rootLogger = logging.getLogger()

consoleHandler = logging.StreamHandler()
consoleHandler.setFormatter(logFormatter)
rootLogger.addHandler(consoleHandler)

settings = json.load(open(settings_path, mode='r'))

database_user = settings['credentials']['db_user']
database_password = settings['credentials']['password']
hostname = settings['hostname']
schema_name = settings['schema']

weeks_to_load = settings['weeks_no']
year_to_load = settings['year_no']

employees = settings['employees']

send_mail_to = settings['send_mail_to']

logging.info(f'Settings were loaded from file: {settings_path}')
logging.info(f'The following periods were specified: {", ".join(list(map(lambda i: "{:02}".format(i), weeks_to_load)))} -> {str(year_to_load)}')
if len(employees) > 0:
    logging.info(f'The following employees were specified: {", ".join(employees)}')

calendar_file = load_workbook(CALENDAR_PATH, data_only=True)
calendar = get_table_data(calendar_file, 'Calendar', 'calendar_actual')
calendar['date'] = calendar['date'].dt.date

logging.info('Calendar has been loaded')

mappings_file = load_workbook(MAPPINGS_PATH, data_only=True)
employees_table = get_table_data(mappings_file, 'Справочники', 'Сотрудники')
all_employees = list(employees_table['Шаблон заголовка'].dropna())

logging.info('List of all employees has been loaded')

try:
    connection = mysql.connector.connect(
        host=hostname,
        database=schema_name,
        user=database_user,
        password=database_password
    )

    logging.info(f'[MYSQL CONNECTION ESTABLISHED]: hostname = {hostname}, database = {schema_name}, user = {database_user}')
    
    cursor = connection.cursor()            

    errors_counter = 0
    total_rows_inserted = 0
    outlook_message_array = []
    errors_messages_array = []

    for week in weeks_to_load:
        week_rows_inserted = 0

        week_folder_name = str(year_to_load) + '-W' + f'{week:02d}'
        current_week_calendar = calendar[
            (calendar['budgetYearNo'] == year_to_load) & 
            (calendar['budgetWeekNo'] == week)
        ]

        week_start_date = current_week_calendar['date'].min()
        week_end_date = current_week_calendar['date'].max()

        logging.info(f'[{week_folder_name}] Week start date: {week_start_date}, week end date: {week_end_date}')

        plans_folder = MASTER_FOLDER + '\\План\\' + week_folder_name
        is_folder_exists = os.path.isdir(plans_folder)
        plans = os.listdir(plans_folder) if is_folder_exists else []

        logging.info(f'[{week_folder_name}] Plans folder: {plans_folder}, folder exists: {is_folder_exists} ')
        logging.info(f'[{week_folder_name}] The following items ({len(plans)}) are in the folder:{(NEW_LINE+TAB)}{(NEW_LINE+TAB).join(plans)}')

        for file_path in plans:
            try:            
                file_name_pattern = f'plan_{week_folder_name}_'
                extension = '.xlsm'
                employee = file_path[
                    file_path.find(file_name_pattern) + len(file_name_pattern):
                    file_path.find(extension)
                ]

                if file_path.startswith(file_name_pattern) and (True if len(employees) == 0 else employee in employees) and (employee in all_employees) and file_path.endswith(extension):
                    plan_path = f'{plans_folder}\\{file_path}' 

                    plan_file = load_workbook(plan_path, data_only=True)

                    # fnGetPlan

                    # get planner name
                    planner_defined_range_description = plan_file.defined_names['PlannerName'].attr_text.split('!')
                    planner_sheet_name = planner_defined_range_description[0].replace("'", "")
                    planner_cell_address = planner_defined_range_description[1].replace('$', '')
                    planner_name = plan_file[planner_sheet_name][planner_cell_address].value

                    source_plan_df = get_table_data(plan_file, 'План CW', 'Plan')
                    source_plan_df['Deadline']= pd.to_datetime(source_plan_df['Deadline']).dt.date

                    logging.info(f'[{week_folder_name}] [{file_path}] The file was loaded to the dataframe')

                    # get plans entries with filled hours
                    plan_df = source_plan_df.drop('Руководитель', axis=1, errors='ignore')
                    plan_df = plan_df.melt(
                        id_vars=[
                            'Проект', 
                            'Задача', 
                            'Исполнитель', 
                            'Deadline', 
                            'Backlog', 
                            '% выполнения'
                        ],
                        var_name='Период',
                        value_name='Часы'
                    )
                    plan_df['Часы'] = pd.to_numeric(plan_df['Часы'], downcast='float')
                    plan_df = plan_df.dropna(subset=['Часы'])

                    df_length = len(plan_df)

                    logging.info(f'[{week_folder_name}] [{file_path}] {"Entries were unpivoted and typed:" : <60} [{str(df_length)}] ')

                    # get backlog plans entries
                    backlog_df = pd.merge(
                        source_plan_df, plan_df, how='outer', 
                        left_on=['Проект', 'Задача', 'Исполнитель', 'Deadline', 'Backlog', '% выполнения'], 
                        right_on=['Проект', 'Задача', 'Исполнитель', 'Deadline', 'Backlog', '% выполнения'], 
                        indicator=True
                    )
                    backlog_df = backlog_df.loc[backlog_df['_merge'] == 'left_only']
                    backlog_df = backlog_df[
                        [
                            'Проект', 
                            'Задача',
                            'Исполнитель', 
                            'Deadline',
                            'Backlog',
                            '% выполнения'
                        ]
                    ]

                    logging.info(f'[{week_folder_name}] [{file_path}] {"Backlog tasks were selected:" : <60} [{str(len(backlog_df))}]')

                    # combine all entries
                    plan_result_df = pd.concat([plan_df, backlog_df])

                    df_length_before_filter = len(plan_result_df)
                    logging.info(f'[{week_folder_name}] [{file_path}] {"Backlog was appended:" : <60} [{str(df_length_before_filter)}]')

                    # filter out invalid rows
                    # plan_result_df = plan_result_df[
                    #     plan_result_df.apply(
                    #         lambda x: (not pd.isnull(x['Проект'])) and 
                    #         (not pd.isnull(x['Исполнитель'])) and 
                    #         ( (not pd.isnull(x['Часы'])) or (not pd.isnull(x['Backlog'])) ), 
                    #         axis=1
                    #     )
                    # ]

                    plan_result_df = plan_result_df[
                        (plan_result_df['Проект'].notnull()) &
                        (plan_result_df['Исполнитель'].notnull()) & 
                        ( (plan_result_df['Часы'].notnull()) | (plan_result_df['Backlog'].notnull()) )
                    ]

                    df_length_after_filter = len(plan_result_df)

                    logging.info(f'[{week_folder_name}] [{file_path}] {"`Null` rows were filtered out:" : <60} [{df_length_after_filter}]')

                    plan_result_df['Постановщик'] = planner_name
                    plan_result_df['Дата плана'] = week_start_date
                    
                    plan_result_df = plan_result_df[
                        (plan_result_df['Часы'] != 0) | 
                        (plan_result_df['Backlog'] == 'Backlog')
                    ]

                    if len(plan_result_df) > 0:
                        plan_result_df = plan_result_df[[
                            'Дата плана', 
                            'Постановщик', 
                            'Проект', 
                            'Задача', 
                            'Исполнитель', 
                            'Deadline',
                            'Backlog',
                            'Период',
                            'Часы',
                            '% выполнения'
                        ]]

                        plan_result_df = plan_result_df.replace({np.nan: None})

                        entries_to_load = list(map(tuple, plan_result_df.itertuples(index=False)))
                    
                        logging.info(f'[{week_folder_name}] [{file_path}] {"Dataframe was converted to a list of tuples:" : <60} [{len(entries_to_load)}]')

                        DELETE_STATEMENT = f"DELETE FROM plans WHERE plan_date = '{week_start_date}' and planner = '{planner_name}'"

                        cursor.execute(DELETE_STATEMENT)
                        connection.commit()

                        logging.critical(f'[{week_folder_name}] [{file_path}] {"Rows were deleted from the database" : <60} {"[" + str(cursor.rowcount)+ "]": >5}')

                        cursor.executemany(INSERT_STATEMENT, entries_to_load)  
                        connection.commit()
                        
                        result_message = f'[{week_folder_name}] [{file_path}] {"Rows were inserted to the database" : <60} {"[" + str(cursor.rowcount)+ "]": >5}'
                        total_rows_inserted += cursor.rowcount
                        week_rows_inserted += cursor.rowcount
                        logging.critical(result_message)

                        outlook_message_array.append(result_message)
                    else:
                        result_message = f'[{week_folder_name}] [{file_path}] File is empty or didn`t pass validation'
                        logging.critical(result_message)
                        outlook_message_array.append(result_message)
            
            except Exception as e:
                result_message = f'[{week_folder_name}] [{file_path}] Exception was raised:{NEW_LINE}{e}'
                errors_counter += 1
                errors_messages_array.append(result_message)
                logging.error(result_message)
        
        week_result_message = f'[{week_folder_name}] {"Total number of inserted rows:" : <60} [{week_rows_inserted}]{NEW_LINE}'
        outlook_message_array.append(week_result_message)
        logging.info(week_result_message)

except mysql.connector.Error as error:
    result_message = f'Exception was raised:{NEW_LINE}{error}'
    errors_counter += 1
    errors_messages_array.append(result_message)
    logging.error(result_message)

finally:
    if connection.is_connected():
        cursor.close()
        connection.close()
        logging.info('MySQL connection is closed')

refresh_end_time = datetime.datetime.now()
refresh_duration = refresh_end_time - refresh_start_time

logging.info(f'Refresh took {refresh_duration.total_seconds()} seconds')

outlook = win32.Dispatch('outlook.application')
mail = outlook.CreateItem(0)
mail.To = ';'.join(send_mail_to)
mail.Subject = 'Plans refresh result'

mail_message = f'''{'SUCCESS' if errors_counter == 0 else 'FAILURE'}

Plans refresh took: {refresh_duration.total_seconds()} seconds.
Total number of loaded rows: {str(total_rows_inserted)}.
Total number of errors raised: {str(errors_counter)}.

Logs file could be seen in the attachements.
The logs file could be found at: {logs_file_name}

Detailed info:
{NEW_LINE.join([message for message in outlook_message_array])}
{(NEW_LINE + 'Detailed info about errors:' + NEW_LINE + NEW_LINE.join([error for error in errors_messages_array]) if errors_counter > 0 else '')}
'''

mail.Body = mail_message
mail.Attachments.Add(logs_file_name)
mail.Send()